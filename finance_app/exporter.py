from __future__ import annotations

from collections import defaultdict
from datetime import datetime
from pathlib import Path
import calendar
import zipfile
from xml.sax.saxutils import escape


MONTH_NAMES = {
    1: "Enero",
    2: "Febrero",
    3: "Marzo",
    4: "Abril",
    5: "Mayo",
    6: "Junio",
    7: "Julio",
    8: "Agosto",
    9: "Septiembre",
    10: "Octubre",
    11: "Noviembre",
    12: "Diciembre",
}

COLORS = {
    "header_bg": "D8E2F1",
    "header_text": "1F2937",
    "ingreso_bg": "E9F7EF",
    "ingreso_text": "136A3A",
    "gasto_bg": "FDECEC",
    "gasto_text": "9F1D1D",
    "total_bg": "FFF3E0",
    "total_text": "7A3E00",
    "balance_pos_bg": "E8F5E9",
    "balance_neg_bg": "FDECEA",
    "balance_zero_bg": "EDF0F4",
    "dashboard_blue": "E8EEF8",
    "dashboard_green": "E9F7EF",
    "dashboard_red": "FDECEC",
    "dashboard_orange": "FFF3E0",
}


def export_to_xlsx(rows: list[dict], output_path: Path) -> Path:
    """Backward-compatible export entrypoint used by filtered/detail exports.

    If openpyxl is available, uses formatted workbook. Otherwise falls back to
    lightweight internal XLSX writer.
    """
    output_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Movimientos"
        _write_movimientos_sheet(ws, rows)
        _auto_width(ws)
        wb.save(output_path)
        return output_path
    except Exception:
        return _export_basic_xlsx(rows, output_path)


def export_filtered_movimientos(rows: list[dict], output_path: Path) -> Path:
    return export_to_xlsx(rows, output_path)


def export_monthly_report(service, month: int, year: int, output_path: Path) -> Path:
    rows = service.list_movimientos(month, year)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        from openpyxl import Workbook

        wb = Workbook()
        wb.remove(wb.active)

        _add_month_summary_sheet(wb, service, rows, month, year)

        ws_mov = wb.create_sheet("Movimientos")
        _write_movimientos_sheet(ws_mov, rows)

        income_rows = [r for r in rows if r.get("tipo") == "ingreso"]
        expense_rows = [r for r in rows if r.get("tipo") == "gasto"]
        savings_rows = [r for r in rows if r.get("tipo") == "ahorro"]
        investment_rows = [r for r in rows if r.get("tipo") == "inversion" or "invers" in str(r.get("categoria", "")).lower()]

        ws_income = wb.create_sheet("Ingresos")
        _write_type_sheet(ws_income, income_rows, "ingreso")

        ws_expense = wb.create_sheet("Gastos")
        _write_type_sheet(ws_expense, expense_rows, "gasto")

        ws_savings = wb.create_sheet("Ahorros")
        _write_type_sheet(ws_savings, savings_rows, "ingreso")

        ws_investments = wb.create_sheet("Inversiones")
        _write_type_sheet(ws_investments, investment_rows, "ingreso")

        ws_by_cat = wb.create_sheet("Gastos por categoria")
        _write_expenses_by_category_sheet(ws_by_cat, service.get_expenses_by_category(month, year, "gasto"))

        ws_categories = wb.create_sheet("Categorias")
        _write_categories_sheet(ws_categories, service.list_categorias())

        ws_fixed = wb.create_sheet("Gastos fijos")
        _write_fixed_expenses_sheet(ws_fixed, service.list_gastos_fijos())

        ws_budgets = wb.create_sheet("Presupuestos")
        _write_budgets_sheet(ws_budgets, service.list_presupuestos(month, year))

        ws_goals = wb.create_sheet("Metas")
        _write_goals_sheet(ws_goals, service.list_metas_ahorro())

        plan_rows = _pending_plan_for_month(service, month, year)
        if plan_rows:
            ws_plan = wb.create_sheet("Planificacion")
            _write_plan_sheet(ws_plan, plan_rows)

        wb.save(output_path)
        return output_path
    except Exception:
        return _export_basic_xlsx(rows, output_path)


def export_date_range_report(service, desde: str, hasta: str, output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    rows = _list_movimientos_between(service, desde, hasta)
    try:
        from openpyxl import Workbook

        wb = Workbook()
        wb.remove(wb.active)

        ws_summary = wb.create_sheet("Resumen")
        totals = _totals_for_rows(rows)
        ws_summary.append(["Campo", "Valor"])
        _style_header(ws_summary, 1)
        summary_items = [
            ("Período", f"{desde} a {hasta}"),
            ("Total ingresos", totals["ingreso"]),
            ("Total gastos", totals["gasto"]),
            ("Total ahorro", totals["ahorro"]),
            ("Total inversión", totals["inversion"]),
            ("Balance", totals["balance"]),
            ("Cantidad de movimientos", len(rows)),
            ("Fecha de exportación", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ]
        for label, value in summary_items:
            ws_summary.append([label, value])
        _style_table(ws_summary, money_cols={2}, text_left_cols={1})
        _auto_width(ws_summary)

        ws_mov = wb.create_sheet("Movimientos")
        _write_movimientos_sheet(ws_mov, rows)

        ws_income = wb.create_sheet("Ingresos")
        _write_type_sheet(ws_income, [r for r in rows if r.get("tipo") == "ingreso"], "ingreso")

        ws_expense = wb.create_sheet("Gastos")
        _write_type_sheet(ws_expense, [r for r in rows if r.get("tipo") == "gasto"], "gasto")

        ws_savings = wb.create_sheet("Ahorros")
        _write_type_sheet(ws_savings, [r for r in rows if r.get("tipo") == "ahorro"], "ingreso")

        ws_investments = wb.create_sheet("Inversiones")
        _write_type_sheet(ws_investments, [r for r in rows if r.get("tipo") == "inversion"], "ingreso")

        wb.save(output_path)
        return output_path
    except Exception:
        return _export_basic_xlsx(rows, output_path)


def export_yearly_report(service, year: int, output_path: Path) -> Path:
    rows = service.list_movimientos_by_year(year)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        from openpyxl import Workbook

        wb = Workbook()
        wb.remove(wb.active)

        _add_year_summary_sheet(wb, service, rows, year)

        ws_mov = wb.create_sheet("Movimientos del ano")
        _write_movimientos_sheet(ws_mov, rows)

        ws_month = wb.create_sheet("Resumen por mes")
        _write_month_balance_sheet(ws_month, service, year)

        income_rows = [r for r in rows if r.get("tipo") == "ingreso"]
        expense_rows = [r for r in rows if r.get("tipo") == "gasto"]
        savings_rows = [r for r in rows if r.get("tipo") == "ahorro"]
        investment_rows = [r for r in rows if r.get("tipo") == "inversion" or "invers" in str(r.get("categoria", "")).lower()]

        ws_income = wb.create_sheet("Ingresos")
        _write_type_sheet(ws_income, income_rows, "ingreso")

        ws_expense = wb.create_sheet("Gastos")
        _write_type_sheet(ws_expense, expense_rows, "gasto")

        ws_savings = wb.create_sheet("Ahorros")
        _write_type_sheet(ws_savings, savings_rows, "ingreso")

        ws_investments = wb.create_sheet("Inversiones")
        _write_type_sheet(ws_investments, investment_rows, "ingreso")

        ws_by_cat = wb.create_sheet("Gastos por categoria")
        _write_expenses_by_category_sheet(ws_by_cat, _expense_categories_for_year(rows))

        ws_balance = wb.create_sheet("Balance mensual")
        _write_month_balance_sheet(ws_balance, service, year)

        plan_rows = _pending_plan_for_year(service, year)
        if plan_rows:
            ws_plan = wb.create_sheet("Planificacion")
            _write_plan_sheet(ws_plan, plan_rows)

        _add_charts_to_year_summary(wb, service, year, rows)

        wb.save(output_path)
        return output_path
    except Exception:
        return _export_basic_xlsx(rows, output_path)


def _write_movimientos_sheet(ws, rows: list[dict]) -> None:
    headers = ["Fecha", "Dia", "Tipo", "Categoria", "Descripcion", "Monto", "Saldo acumulado"]
    ws.append(headers)
    _style_header(ws, 1)

    for row in rows:
        weekday = _weekday_name(int(row.get("dia_semana_num", _weekday_from_date(row.get("fecha", "")))))
        amount = float(row.get("monto", 0.0) or 0.0)
        saldo = float(row.get("saldo_acumulado", 0.0) or 0.0)
        ws.append(
            [
                str(row.get("fecha", "")),
                weekday,
                str(row.get("tipo", "")),
                str(row.get("categoria", "")),
                str(row.get("descripcion", "")),
                amount,
                saldo,
            ]
        )

    _style_table(ws, money_cols={6, 7}, text_left_cols={4, 5})
    _freeze_and_filter(ws)

    for i in range(2, ws.max_row + 1):
        move_type = str(ws.cell(i, 3).value or "")
        if move_type == "ingreso":
            _style_movement_row(ws, i, "ingreso")
        elif move_type == "gasto":
            _style_movement_row(ws, i, "gasto")
        _style_balance_cell(ws.cell(i, 7))

    _auto_width(ws)


def _write_type_sheet(ws, rows: list[dict], move_type: str) -> None:
    headers = ["Fecha", "Categoria", "Descripcion", "Monto"]
    ws.append(headers)
    _style_header(ws, 1)

    total = 0.0
    by_category: dict[str, float] = defaultdict(float)
    for row in rows:
        amount = float(row.get("monto", 0.0) or 0.0)
        total += amount
        category = str(row.get("categoria", ""))
        by_category[category] += amount
        ws.append([
            str(row.get("fecha", "")),
            category,
            str(row.get("descripcion", "")),
            amount,
        ])
        _style_movement_row(ws, ws.max_row, move_type, cols=(1, 2, 3, 4))

    ws.append(["", "", "TOTAL", total])
    _highlight_total_row(ws, ws.max_row)

    start = ws.max_row + 2
    ws.cell(start, 1, "Resumen por categoria")
    ws.cell(start + 1, 1, "Categoria")
    ws.cell(start + 1, 2, "Total")
    _style_header(ws, start + 1)

    r = start + 2
    for category, amount in sorted(by_category.items(), key=lambda item: item[1], reverse=True):
        ws.cell(r, 1, category)
        ws.cell(r, 2, amount)
        r += 1

    if r > start + 2:
        _style_range_table(ws, start + 1, r - 1, money_cols={2}, text_left_cols={1})

    _style_table(ws, money_cols={4}, text_left_cols={2, 3})
    _freeze_and_filter(ws)
    _auto_width(ws)


def _write_expenses_by_category_sheet(ws, rows: list[dict]) -> None:
    headers = ["Categoria", "Total gastado", "Cantidad de movimientos", "% sobre gastos"]
    ws.append(headers)
    _style_header(ws, 1)

    total_expenses = sum(float(r.get("total", 0.0) or 0.0) for r in rows)
    for row in rows:
        amount = float(row.get("total", 0.0) or 0.0)
        count = int(row.get("movimientos", 0) or 0)
        pct = (amount / total_expenses) if total_expenses > 0 else 0.0
        ws.append([str(row.get("categoria", "")), amount, count, pct])

    ws.append(["TOTAL", total_expenses, sum(int(r.get("movimientos", 0) or 0) for r in rows), 1.0 if total_expenses > 0 else 0.0])
    _highlight_total_row(ws, ws.max_row)
    if rows:
        _highlight_top_category_row(ws, 2)
        _apply_data_bar(ws, 2, max(2, ws.max_row - 1))

    _style_table(ws, money_cols={2}, percent_cols={4}, text_left_cols={1})
    _freeze_and_filter(ws)
    _auto_width(ws)


def _write_plan_sheet(ws, rows: list[dict]) -> None:
    headers = ["Vencimiento", "Estado", "Categoria", "Descripcion", "Monto estimado", "Recurrente", "Frecuencia"]
    ws.append(headers)
    _style_header(ws, 1)

    total = 0.0
    for row in rows:
        amount = float(row.get("monto_estimado", 0.0) or 0.0)
        total += amount
        ws.append([
            str(row.get("fecha_vencimiento", "")),
            str(row.get("estado", "")),
            str(row.get("categoria", "")),
            str(row.get("descripcion", "")),
            amount,
            "Si" if int(row.get("es_recurrente", 0) or 0) == 1 else "No",
            str(row.get("frecuencia", "") or ""),
        ])

    ws.append(["", "", "", "TOTAL", total, "", ""])
    _highlight_total_row(ws, ws.max_row)

    _style_table(ws, money_cols={5}, text_left_cols={3, 4, 7})
    _freeze_and_filter(ws)
    _auto_width(ws)


def _write_categories_sheet(ws, rows: list[dict]) -> None:
    ws.append(["Nombre", "Tipo"])
    _style_header(ws, 1)
    for row in rows:
        ws.append([str(row.get("nombre", "")), str(row.get("tipo", ""))])
    _style_table(ws, text_left_cols={1, 2})
    _freeze_and_filter(ws)
    _auto_width(ws)


def _write_fixed_expenses_sheet(ws, rows: list[dict]) -> None:
    ws.append(["Categoria", "Descripcion", "Monto", "Dia vencimiento", "Activo"])
    _style_header(ws, 1)
    for row in rows:
        ws.append([
            str(row.get("categoria", "")),
            str(row.get("descripcion", "")),
            float(row.get("monto", 0.0) or 0.0),
            int(row.get("dia_vencimiento", 0) or 0),
            "Si" if int(row.get("activo", 0) or 0) == 1 else "No",
        ])
    _style_table(ws, money_cols={3}, text_left_cols={1, 2, 5})
    _freeze_and_filter(ws)
    _auto_width(ws)


def _write_budgets_sheet(ws, rows: list[dict]) -> None:
    ws.append(["Categoria", "Mes", "Ano", "Presupuesto", "Gastado", "Disponible", "Excedido"])
    _style_header(ws, 1)
    for row in rows:
        month = int(row.get("mes", 0) or 0)
        ws.append([
            str(row.get("categoria", "")),
            _month_label(month),
            int(row.get("anio", 0) or 0),
            float(row.get("monto_presupuestado", 0.0) or 0.0),
            float(row.get("monto_gastado", 0.0) or 0.0),
            float(row.get("monto_disponible", 0.0) or 0.0),
            "Si" if bool(row.get("excedido")) else "No",
        ])
    _style_table(ws, money_cols={4, 5, 6}, text_left_cols={1, 2, 7})
    _freeze_and_filter(ws)
    _auto_width(ws)


def _write_goals_sheet(ws, rows: list[dict]) -> None:
    ws.append(["Nombre", "Objetivo", "Ahorrado", "Faltante", "Estado"])
    _style_header(ws, 1)
    for row in rows:
        ws.append([
            str(row.get("nombre", "")),
            float(row.get("monto_objetivo", 0.0) or 0.0),
            float(row.get("monto_ahorrado", 0.0) or 0.0),
            float(row.get("faltante", 0.0) or 0.0),
            str(row.get("estado", "")),
        ])
    _style_table(ws, money_cols={2, 3, 4}, text_left_cols={1, 5})
    _freeze_and_filter(ws)
    _auto_width(ws)


def _add_month_summary_sheet(wb, service, rows: list[dict], month: int, year: int) -> None:
    ws = wb.create_sheet("Resumen")
    totals = service.get_resumen_mensual_con_saldo(month, year)
    income_count = sum(1 for r in rows if r.get("tipo") == "ingreso")
    expense_count = sum(1 for r in rows if r.get("tipo") == "gasto")
    savings_count = sum(1 for r in rows if r.get("tipo") == "ahorro")

    by_cat = service.get_expenses_by_category(month, year, "gasto")
    top_cat = by_cat[0]["categoria"] if by_cat else "-"

    pending = _pending_plan_for_month(service, month, year)
    pending_total = sum(float(r.get("monto_estimado", 0.0) or 0.0) for r in pending)

    items = [
        ("Mes/Ano", f"{MONTH_NAMES.get(month, month)} {year}"),
        ("Saldo inicial", float(totals["saldo_inicial"])),
        ("Total ingresos", float(totals["ingreso"])),
        ("Total gastos", float(totals["gasto"])),
        ("Total ahorro", float(totals.get("ahorro", 0.0))),
        ("Balance final", float(totals["balance_final"])),
        ("Disponible luego de ahorro", float(totals.get("disponible_luego_ahorro", 0.0))),
        ("Cantidad de ingresos", income_count),
        ("Cantidad de gastos", expense_count),
        ("Cantidad de ahorros", savings_count),
        ("Categoria con mayor gasto", top_cat),
        ("Gastos programados pendientes del mes", pending_total),
    ]
    _write_kv_summary(ws, items)


def _add_year_summary_sheet(wb, service, rows: list[dict], year: int) -> None:
    ws = wb.create_sheet("Resumen anual")

    total_income = sum(float(r.get("monto", 0.0) or 0.0) for r in rows if r.get("tipo") == "ingreso")
    total_expense = sum(float(r.get("monto", 0.0) or 0.0) for r in rows if r.get("tipo") == "gasto")

    by_month_income = defaultdict(float)
    by_month_expense = defaultdict(float)
    by_cat_expense = defaultdict(float)
    for r in rows:
        date_str = str(r.get("fecha", ""))
        month = int(date_str[5:7]) if len(date_str) >= 7 and date_str[5:7].isdigit() else 0
        amount = float(r.get("monto", 0.0) or 0.0)
        if r.get("tipo") == "ingreso":
            by_month_income[month] += amount
        elif r.get("tipo") == "gasto":
            by_month_expense[month] += amount
            by_cat_expense[str(r.get("categoria", ""))] += amount

    max_exp_month = _month_label(max(by_month_expense, key=by_month_expense.get)) if by_month_expense else "-"
    max_inc_month = _month_label(max(by_month_income, key=by_month_income.get)) if by_month_income else "-"
    top_cat = max(by_cat_expense, key=by_cat_expense.get) if by_cat_expense else "-"

    items = [
        ("Ano", year),
        ("Total ingresos anual", total_income),
        ("Total gastos anual", total_expense),
        ("Balance anual", total_income - total_expense),
        ("Mes con mas gastos", max_exp_month),
        ("Mes con mas ingresos", max_inc_month),
        ("Categoria con mayor gasto del ano", top_cat),
        ("Promedio mensual de gastos", total_expense / 12.0),
        ("Promedio mensual de ingresos", total_income / 12.0),
    ]
    _write_kv_summary(ws, items)


def _write_month_balance_sheet(ws, service, year: int) -> None:
    headers = ["Mes", "Saldo inicial", "Ingresos", "Gastos", "Balance final"]
    ws.append(headers)
    _style_header(ws, 1)

    for month in range(1, 13):
        totals = service.get_resumen_mensual_con_saldo(month, year)
        ws.append(
            [
                _month_label(month),
                float(totals["saldo_inicial"]),
                float(totals["ingreso"]),
                float(totals["gasto"]),
                float(totals["balance_final"]),
            ]
        )
        _style_balance_cell(ws.cell(ws.max_row, 5))

    ws.append(["TOTAL", "", "", "", ""])
    _highlight_total_row(ws, ws.max_row)
    _style_table(ws, money_cols={2, 3, 4, 5}, text_left_cols={1})
    _freeze_and_filter(ws)
    _auto_width(ws)


def _add_charts_to_year_summary(wb, service, year: int, rows: list[dict]) -> None:
    try:
        from openpyxl.chart import BarChart, PieChart, Reference
    except Exception:
        return

    if "Resumen por mes" not in wb.sheetnames or "Gastos por categoria" not in wb.sheetnames:
        return

    summary_ws = wb["Resumen anual"]
    month_ws = wb["Resumen por mes"]
    cat_ws = wb["Gastos por categoria"]

    bar = BarChart()
    bar.title = "Ingresos vs Gastos por mes"
    bar.y_axis.title = "Monto"
    bar.x_axis.title = "Mes"
    data = Reference(month_ws, min_col=3, max_col=4, min_row=1, max_row=13)
    cats = Reference(month_ws, min_col=1, min_row=2, max_row=13)
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(cats)
    bar.height = 6
    bar.width = 10
    summary_ws.add_chart(bar, "D2")

    if cat_ws.max_row > 2:
        pie = PieChart()
        pie.title = "Gastos por categoria"
        data2 = Reference(cat_ws, min_col=2, min_row=1, max_row=cat_ws.max_row - 1)
        labels2 = Reference(cat_ws, min_col=1, min_row=2, max_row=cat_ws.max_row - 1)
        pie.add_data(data2, titles_from_data=True)
        pie.set_categories(labels2)
        pie.height = 6
        pie.width = 8
        summary_ws.add_chart(pie, "D20")


def _write_kv_summary(ws, items: list[tuple[str, object]]) -> None:
    ws.append(["Campo", "Valor"])
    _style_header(ws, 1)
    for label, value in items:
        ws.append([label, value])

    money_labels = {
        "Saldo inicial",
        "Total ingresos",
        "Total gastos",
        "Balance final",
        "Total ahorro",
        "Disponible luego de ahorro",
        "Gastos programados pendientes del mes",
        "Total ingresos anual",
        "Total gastos anual",
        "Balance anual",
        "Promedio mensual de gastos",
        "Promedio mensual de ingresos",
    }

    for r in range(2, ws.max_row + 1):
        label = str(ws.cell(r, 1).value or "")
        if label in money_labels:
            ws.cell(r, 2).number_format = "$ #,##0.00"
            ws.cell(r, 2).alignment = _align_right()
        else:
            ws.cell(r, 2).alignment = _align_left()

        _style_summary_card_row(ws, r, label)
        if label in {
            "Total ingresos",
            "Total gastos",
            "Balance final",
            "Total ingresos anual",
            "Total gastos anual",
            "Balance anual",
            "Promedio mensual de gastos",
            "Promedio mensual de ingresos",
        }:
            _highlight_total_row(ws, r)
        if label in {"Balance final", "Balance anual"}:
            _style_balance_cell(ws.cell(r, 2))

    _style_table(ws, start_row=1, end_row=ws.max_row, start_col=1, end_col=2, text_left_cols={1})
    ws.freeze_panes = "A2"
    _auto_width(ws)


def _pending_plan_for_month(service, month: int, year: int) -> list[dict]:
    rows = service.get_gastos_programados("pendiente")
    out = []
    for row in rows:
        due = str(row.get("fecha_vencimiento", ""))
        if len(due) >= 7 and due[:4].isdigit() and due[5:7].isdigit():
            if int(due[:4]) == year and int(due[5:7]) == month:
                out.append(row)
    return out


def _pending_plan_for_year(service, year: int) -> list[dict]:
    rows = service.get_gastos_programados("pendiente")
    return [r for r in rows if str(r.get("fecha_vencimiento", "")).startswith(f"{year}-")]


def _list_movimientos_between(service, desde: str, hasta: str) -> list[dict]:
    with service.db.connect() as conn:
        rows = conn.execute(
            """
            SELECT
                m.id,
                m.fecha,
                m.tipo,
                c.nombre AS categoria,
                m.descripcion,
                m.monto,
                CAST(strftime('%w', m.fecha) AS INTEGER) AS dia_semana_num,
                (
                    SELECT COALESCE(SUM(
                        CASE
                            WHEN m2.tipo = 'ingreso' THEN m2.monto
                            ELSE -m2.monto
                        END
                    ), 0)
                    FROM movimientos m2
                    WHERE m2.fecha < m.fecha OR (m2.fecha = m.fecha AND m2.id <= m.id)
                ) AS saldo_acumulado
            FROM movimientos m
            JOIN categorias c ON c.id = m.categoria_id
            WHERE m.fecha >= ? AND m.fecha <= ?
            ORDER BY m.fecha DESC, m.id DESC
            """,
            (desde, hasta),
        ).fetchall()
    return [dict(row) for row in rows]


def _totals_for_rows(rows: list[dict]) -> dict[str, float]:
    ingresos = sum(float(r.get("monto", 0) or 0) for r in rows if r.get("tipo") == "ingreso")
    gastos = sum(float(r.get("monto", 0) or 0) for r in rows if r.get("tipo") == "gasto")
    ahorro = sum(float(r.get("monto", 0) or 0) for r in rows if r.get("tipo") == "ahorro")
    inversion = sum(float(r.get("monto", 0) or 0) for r in rows if r.get("tipo") == "inversion")
    return {
        "ingreso": ingresos,
        "gasto": gastos,
        "ahorro": ahorro,
        "inversion": inversion,
        "balance": ingresos - gastos - ahorro - inversion,
    }


def _expense_categories_for_year(rows: list[dict]) -> list[dict]:
    by_cat: dict[str, dict] = {}
    for row in rows:
        if row.get("tipo") != "gasto":
            continue
        cat = str(row.get("categoria", ""))
        amount = float(row.get("monto", 0.0) or 0.0)
        if cat not in by_cat:
            by_cat[cat] = {"categoria": cat, "total": 0.0, "movimientos": 0}
        by_cat[cat]["total"] += amount
        by_cat[cat]["movimientos"] += 1
    return sorted(by_cat.values(), key=lambda x: x["total"], reverse=True)


def _freeze_and_filter(ws) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{_col_name(ws.max_column)}{max(1, ws.max_row)}"


def _style_table(ws, money_cols: set[int] | None = None, percent_cols: set[int] | None = None, text_left_cols: set[int] | None = None,
                 start_row: int = 2, end_row: int | None = None, start_col: int = 1, end_col: int | None = None) -> None:
    money_cols = money_cols or set()
    percent_cols = percent_cols or set()
    text_left_cols = text_left_cols or set()
    end_row = end_row or ws.max_row
    end_col = end_col or ws.max_column

    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            cell = ws.cell(r, c)
            cell.border = _border()
            if c in money_cols and isinstance(cell.value, (int, float)):
                cell.number_format = "$ #,##0.00"
                cell.alignment = _align_right()
            elif c in percent_cols and isinstance(cell.value, (int, float)):
                cell.number_format = "0.00%"
                cell.alignment = _align_right()
            elif c in text_left_cols:
                cell.alignment = _align_left()
            elif isinstance(cell.value, (int, float)):
                cell.alignment = _align_right()
            else:
                cell.alignment = _align_left()


def _style_range_table(ws, start_row: int, end_row: int, money_cols: set[int] | None = None, text_left_cols: set[int] | None = None) -> None:
    _style_table(ws, money_cols=money_cols or set(), text_left_cols=text_left_cols or set(), start_row=start_row, end_row=end_row)


def _style_header(ws, row_idx: int) -> None:
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row_idx, c)
        cell.font = _font(bold=True, color=COLORS["header_text"])
        cell.fill = _fill(COLORS["header_bg"])
        cell.border = _border()
        cell.alignment = _align_center()
    ws.row_dimensions[row_idx].height = 22


def _highlight_total_row(ws, row_idx: int) -> None:
    from openpyxl.styles import Border, Side

    top = Side(style="medium", color="C97B2D")
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row_idx, c)
        cell.font = _font(bold=True, color=COLORS["total_text"])
        cell.fill = _fill(COLORS["total_bg"])
        base = _border()
        cell.border = Border(left=base.left, right=base.right, top=top, bottom=base.bottom)


def _auto_width(ws) -> None:
    for col in range(1, ws.max_column + 1):
        letter = _col_name(col)
        max_len = 0
        has_money_format = False
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row, col)
            value = cell.value
            txt = "" if value is None else str(value)
            max_len = max(max_len, len(txt))
            if isinstance(cell.number_format, str) and "$" in cell.number_format:
                has_money_format = True
        base_width = max_len + 3
        if has_money_format:
            base_width = max(base_width, 16)
        ws.column_dimensions[letter].width = min(max(10, base_width), 48)


def _style_movement_row(ws, row_idx: int, move_type: str, cols: tuple[int, ...] | None = None) -> None:
    cols = cols or tuple(range(1, ws.max_column + 1))
    if move_type == "ingreso":
        fill_color = COLORS["ingreso_bg"]
        text_color = COLORS["ingreso_text"]
        border_color = "9ED7B5"
    else:
        fill_color = COLORS["gasto_bg"]
        text_color = COLORS["gasto_text"]
        border_color = "F0B5B5"

    for c in cols:
        cell = ws.cell(row_idx, c)
        cell.fill = _fill(fill_color)
        cell.border = _border(border_color)
        if c in {3, 4, 6}:
            cell.font = _font(color=text_color)


def _style_balance_cell(cell) -> None:
    value = cell.value
    if not isinstance(value, (int, float)):
        return
    if value > 0:
        cell.fill = _fill(COLORS["balance_pos_bg"])
        cell.font = _font(bold=True, color=COLORS["ingreso_text"])
    elif value < 0:
        cell.fill = _fill(COLORS["balance_neg_bg"])
        cell.font = _font(bold=True, color=COLORS["gasto_text"])
    else:
        cell.fill = _fill(COLORS["balance_zero_bg"])
        cell.font = _font(bold=True, color=COLORS["header_text"])


def _style_summary_card_row(ws, row_idx: int, label: str) -> None:
    if label in {"Saldo inicial", "Mes/Ano", "Ano"}:
        color = COLORS["dashboard_blue"]
    elif "ingresos" in label.lower():
        color = COLORS["dashboard_green"]
    elif "gastos" in label.lower():
        color = COLORS["dashboard_red"]
    elif "balance" in label.lower():
        color = COLORS["dashboard_orange"]
    else:
        color = "FFFFFF"

    ws.cell(row_idx, 1).fill = _fill(color)
    ws.cell(row_idx, 2).fill = _fill(color)
    ws.cell(row_idx, 1).font = _font(bold=True)


def _highlight_top_category_row(ws, row_idx: int) -> None:
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row_idx, c)
        cell.fill = _fill("FFF7D6")
        cell.font = _font(bold=True, color="7A3E00")


def _apply_data_bar(ws, start_row: int, end_row: int) -> None:
    try:
        from openpyxl.formatting.rule import DataBarRule
    except Exception:
        return
    if end_row < start_row:
        return
    rule = DataBarRule(start_type="num", start_value=0, end_type="max", color="F4B183", showValue=True)
    ws.conditional_formatting.add(f"B{start_row}:B{end_row}", rule)


def _font(bold: bool = False, color: str = "1F2937"):
    from openpyxl.styles import Font

    return Font(name="Segoe UI", sz=10, bold=bold, color=color)


def _fill(color: str = "E8EEF8"):
    from openpyxl.styles import PatternFill

    return PatternFill(fill_type="solid", start_color=color, end_color=color)


def _border(color: str = "D0D7E2"):
    from openpyxl.styles import Border, Side

    side = Side(style="thin", color=color)
    return Border(left=side, right=side, top=side, bottom=side)


def _align_left():
    from openpyxl.styles import Alignment

    return Alignment(horizontal="left", vertical="center")


def _align_right():
    from openpyxl.styles import Alignment

    return Alignment(horizontal="right", vertical="center")


def _align_center():
    from openpyxl.styles import Alignment

    return Alignment(horizontal="center", vertical="center")


def _month_label(month: int) -> str:
    return MONTH_NAMES.get(month, str(month))


def _weekday_from_date(date_str: str) -> int:
    try:
        dt = datetime.strptime(str(date_str), "%Y-%m-%d")
        # Align to existing mapping where Sunday=0
        return (dt.weekday() + 1) % 7
    except Exception:
        return 0


def _export_basic_xlsx(rows: list[dict], output_path: Path) -> Path:
    has_extended_cols = any("saldo_acumulado" in row or "dia_semana_num" in row for row in rows)
    if has_extended_cols:
        headers = ["Fecha", "Dia", "Tipo", "Categoria", "Descripcion", "Monto", "Saldo acumulado"]
        body_rows = [
            [
                str(r.get("fecha", "")),
                _weekday_name(int(r.get("dia_semana_num", 0))),
                str(r.get("tipo", "")),
                str(r.get("categoria", "")),
                str(r.get("descripcion", "")),
                f"{float(r.get('monto', 0)):.2f}",
                f"{float(r.get('saldo_acumulado', 0)):.2f}",
            ]
            for r in rows
        ]
    else:
        headers = ["Fecha", "Tipo", "Categoria", "Descripcion", "Monto"]
        body_rows = [
            [
                str(r.get("fecha", "")),
                str(r.get("tipo", "")),
                str(r.get("categoria", "")),
                str(r.get("descripcion", "")),
                f"{float(r.get('monto', 0)):.2f}",
            ]
            for r in rows
        ]

    with zipfile.ZipFile(output_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("[Content_Types].xml", _content_types_xml())
        zf.writestr("_rels/.rels", _rels_xml())
        zf.writestr("xl/workbook.xml", _workbook_xml())
        zf.writestr("xl/_rels/workbook.xml.rels", _workbook_rels_xml())
        zf.writestr("xl/worksheets/sheet1.xml", _sheet_xml(headers, body_rows))

    return output_path


def _content_types_xml() -> str:
    return """<?xml version=\"1.0\" encoding=\"UTF-8\"?>
<Types xmlns=\"http://schemas.openxmlformats.org/package/2006/content-types\">
  <Default Extension=\"rels\" ContentType=\"application/vnd.openxmlformats-package.relationships+xml\"/>
  <Default Extension=\"xml\" ContentType=\"application/xml\"/>
  <Override PartName=\"/xl/workbook.xml\" ContentType=\"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml\"/>
  <Override PartName=\"/xl/worksheets/sheet1.xml\" ContentType=\"application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml\"/>
</Types>
"""


def _rels_xml() -> str:
    return """<?xml version=\"1.0\" encoding=\"UTF-8\"?>
<Relationships xmlns=\"http://schemas.openxmlformats.org/package/2006/relationships\">
  <Relationship Id=\"rId1\" Type=\"http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument\" Target=\"xl/workbook.xml\"/>
</Relationships>
"""


def _workbook_xml() -> str:
    return """<?xml version=\"1.0\" encoding=\"UTF-8\"?>
<workbook xmlns=\"http://schemas.openxmlformats.org/spreadsheetml/2006/main\" xmlns:r=\"http://schemas.openxmlformats.org/officeDocument/2006/relationships\">
  <sheets>
    <sheet name=\"Movimientos\" sheetId=\"1\" r:id=\"rId1\"/>
  </sheets>
</workbook>
"""


def _workbook_rels_xml() -> str:
    return """<?xml version=\"1.0\" encoding=\"UTF-8\"?>
<Relationships xmlns=\"http://schemas.openxmlformats.org/package/2006/relationships\">
  <Relationship Id=\"rId1\" Type=\"http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet\" Target=\"worksheets/sheet1.xml\"/>
</Relationships>
"""


def _sheet_xml(headers: list[str], rows: list[list[str]]) -> str:
    xml_rows = []
    all_rows = [headers] + rows
    for idx, row in enumerate(all_rows, start=1):
        cells = []
        for cidx, value in enumerate(row, start=1):
            cell_ref = f"{_col_name(cidx)}{idx}"
            cells.append(f'<c r="{cell_ref}" t="inlineStr"><is><t>{escape(value)}</t></is></c>')
        xml_rows.append(f"<row r=\"{idx}\">{''.join(cells)}</row>")

    return (
        "<?xml version=\"1.0\" encoding=\"UTF-8\"?>"
        "<worksheet xmlns=\"http://schemas.openxmlformats.org/spreadsheetml/2006/main\">"
        f"<sheetData>{''.join(xml_rows)}</sheetData>"
        "</worksheet>"
    )


def _col_name(index: int) -> str:
    name = ""
    while index:
        index, rem = divmod(index - 1, 26)
        name = chr(65 + rem) + name
    return name


def _weekday_name(weekday_number: int) -> str:
    names = {
        0: "Dom",
        1: "Lun",
        2: "Mar",
        3: "Mie",
        4: "Jue",
        5: "Vie",
        6: "Sab",
    }
    return names.get(weekday_number, "")
